import tkinter as tk
from tkinter import filedialog, messagebox, simpledialog, ttk
import pandas as pd
import re
from pathlib import Path
import openpyxl
from datetime import datetime
import json
import os
from typing import List, Dict, Tuple, Optional
import logging

CONFIG_FILE = "invoice_config.json"

DEFAULT_CONFIG = {
    "contractor_cell": "R9C6",
    "number_cell": "R2C9",
    "date_cell": "R2C10",
    "items_start_cell": "R20C3",
    "items_cell_text_part": "C",
    "items_cell_numeric_part": "C",
    "items_cell_weight": "R",
    "items_cell_price": "T",
    "show_review_dialog": True
}

class InvoiceProcessor:
    def __init__(self, root):
        self.root = root
        self.root.title("Обработка счетов-фактур")
        self.root.geometry("1000x700")
        self.output_file = None
        self.selected_invoices: List[str] = []
        
        # Setup logging first
        self.setup_logging()
        
        # Then load config (which uses logging)
        self.config = self.load_config()
        
        # Finally create widgets
        self.create_widgets()

    def setup_logging(self):
        # Configure logging
        logging.basicConfig(
            level=logging.INFO,
            format='%(asctime)s - %(levelname)s - %(message)s',
            handlers=[
                logging.StreamHandler()
            ]
        )
        self.logger = logging.getLogger(__name__)

    def log_message(self, message: str):
        """Log a message both to the logger and the GUI text widget"""
        if hasattr(self, 'logger'):
            self.logger.info(message)
        if hasattr(self, 'log_text'):
            self.log_text.insert(tk.END, f"{message}\n")
            self.log_text.see(tk.END)
        else:
            print(message)  # Fallback if GUI not yet initialized

    def create_widgets(self):
        # Main container with padding
        self.main_frame = ttk.Frame(self.root, padding="10")
        self.main_frame.pack(fill=tk.BOTH, expand=True)

        # Top frame for file operations
        self.file_frame = ttk.LabelFrame(self.main_frame, text="Файлы", padding="5")
        self.file_frame.pack(fill=tk.X, pady=(0, 10))

        # Output file selection
        self.output_frame = ttk.Frame(self.file_frame)
        self.output_frame.pack(fill=tk.X, pady=5)
        
        ttk.Label(self.output_frame, text="Файл отчета:").pack(side=tk.LEFT, padx=5)
        self.output_file_label = ttk.Label(self.output_frame, text="Не выбран")
        self.output_file_label.pack(side=tk.LEFT, padx=5)
        ttk.Button(self.output_frame, text="Выбрать", command=self.select_output_file).pack(side=tk.LEFT, padx=5)

        # Invoice files selection
        self.invoice_frame = ttk.Frame(self.file_frame)
        self.invoice_frame.pack(fill=tk.X, pady=5)
        
        ttk.Label(self.invoice_frame, text="Счета-фактуры:").pack(side=tk.LEFT, padx=5)
        ttk.Button(self.invoice_frame, text="Добавить", command=self.select_invoice_files).pack(side=tk.LEFT, padx=5)

        # Settings frame
        self.settings_frame = ttk.LabelFrame(self.main_frame, text="Настройки", padding="5")
        self.settings_frame.pack(fill=tk.X, pady=(0, 10))

        # Show review dialog checkbox
        self.show_dialog_var = tk.BooleanVar(value=self.config.get('show_review_dialog', True))
        self.show_dialog_check = ttk.Checkbutton(
            self.settings_frame, 
            text="Показывать диалог настройки", 
            variable=self.show_dialog_var,
            command=self.save_dialog_setting
        )
        self.show_dialog_check.pack(anchor=tk.W, padx=5, pady=5)

        # Selected invoices list
        self.invoice_list_frame = ttk.LabelFrame(self.main_frame, text="Выбранные счета-фактуры", padding="5")
        self.invoice_list_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 10))

        # Scrollbar for invoice list
        self.scrollbar = ttk.Scrollbar(self.invoice_list_frame)
        self.scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        self.invoice_listbox = tk.Listbox(
            self.invoice_list_frame,
            yscrollcommand=self.scrollbar.set,
            selectmode=tk.EXTENDED
        )
        self.invoice_listbox.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        self.scrollbar.config(command=self.invoice_listbox.yview)

        # Buttons frame
        self.buttons_frame = ttk.Frame(self.main_frame)
        self.buttons_frame.pack(fill=tk.X, pady=(0, 10))

        ttk.Button(
            self.buttons_frame,
            text="Обработать выбранные",
            command=self.process_selected_invoices
        ).pack(side=tk.LEFT, padx=5)

        ttk.Button(
            self.buttons_frame,
            text="Удалить выбранные",
            command=self.remove_selected_invoices
        ).pack(side=tk.LEFT, padx=5)

        # Log frame
        self.log_frame = ttk.LabelFrame(self.main_frame, text="Лог", padding="5")
        self.log_frame.pack(fill=tk.BOTH, expand=True)

        self.log_scrollbar = ttk.Scrollbar(self.log_frame)
        self.log_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        self.log_text = tk.Text(
            self.log_frame,
            height=10,
            yscrollcommand=self.log_scrollbar.set,
            wrap=tk.WORD
        )
        self.log_text.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        self.log_scrollbar.config(command=self.log_text.yview)

    def save_dialog_setting(self):
        self.config['show_review_dialog'] = self.show_dialog_var.get()
        self.save_config()

    def select_output_file(self):
        filename = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if filename:
            self.output_file = filename
            self.output_file_label.config(text=Path(filename).name)
            self.log_message(f"Выбран файл отчета: {filename}")

    def select_invoice_files(self):
        filenames = filedialog.askopenfilenames(filetypes=[("Excel files", "*.xlsx")])
        if filenames:
            for filename in filenames:
                if filename not in self.selected_invoices:
                    self.selected_invoices.append(filename)
            self.update_invoice_listbox()

    def update_invoice_listbox(self):
        self.invoice_listbox.delete(0, tk.END)
        # Sort invoices by number in filename
        sorted_invoices = sorted(
            self.selected_invoices,
            key=lambda x: self.extract_invoice_number_from_filename(x)
        )
        for invoice in sorted_invoices:
            self.invoice_listbox.insert(tk.END, Path(invoice).name)

    def extract_invoice_number_from_filename(self, filename: str) -> str:
        # Extract numbers from filename
        numbers = re.findall(r'\d+', Path(filename).stem)
        return numbers[0] if numbers else ""

    def remove_selected_invoices(self):
        selected_indices = self.invoice_listbox.curselection()
        for index in reversed(selected_indices):
            del self.selected_invoices[index]
        self.update_invoice_listbox()

    def process_selected_invoices(self):
        if not self.output_file:
            messagebox.showerror("Ошибка", "Сначала выберите файл отчета")
            return

        if not self.selected_invoices:
            messagebox.showerror("Ошибка", "Выберите счета-фактуры для обработки")
            return

        try:
            for invoice_file in sorted(
                self.selected_invoices,
                key=lambda x: self.extract_invoice_number_from_filename(x)
            ):
                self.process_single_invoice(invoice_file)
        except Exception as e:
            self.log_message(f"Ошибка при обработке файлов: {str(e)}")
            messagebox.showerror("Ошибка", f"Произошла ошибка при обработке файлов: {str(e)}")

    def process_single_invoice(self, invoice_file: str):
        try:
            invoice_df = pd.read_excel(invoice_file, header=None)
            extracted_data = self.extract_invoice_data(invoice_df)
            
            if self.show_dialog_var.get():
                dialog = DataReviewDialog(self.root, extracted_data, self.config)
                if dialog.result is None:
                    self.log_message(f"Обработка отменена для файла: {invoice_file}")
                    return
                extracted_data, updated_config = dialog.result
                self.config = updated_config
                self.save_config()
            
            # Create DataFrame for new invoice data
            new_rows = []
            for i, item in enumerate(extracted_data['items']):
                row = {
                    0: extracted_data['number']['value'] if i == 0 else '',
                    1: extracted_data['contractor']['value'] if i == 0 else '',
                    2: extracted_data['date']['value'].strftime('%d.%m.%Y') if isinstance(extracted_data['date']['value'], datetime) and i == 0 else str(extracted_data['date']['value']) if i == 0 else '',
                    3: 'Э' if i == 0 else '',
                    4: item['text_part']['value'],
                    5: '',
                    6: item['numeric_part']['value'],
                    7: item['weight']['value'],
                    8: item['price']['value'],
                    9: '',
                    10: '',
                    11: f"{extracted_data['number']['value']} от {extracted_data['date']['value'].strftime('%d.%m.%Y')}" if isinstance(extracted_data['date']['value'], datetime) and i == 0 else f"{extracted_data['number']['value']} от {str(extracted_data['date']['value'])}" if i == 0 else ''
                }
                new_rows.append(row)

            new_df = pd.DataFrame(new_rows)
            self.save_with_formatting(new_df)
            self.log_message(f"Успешно обработан файл: {Path(invoice_file).name}")

        except Exception as e:
            self.log_message(f"Ошибка обработки файла {Path(invoice_file).name}: {str(e)}")
            raise

    def extract_invoice_data(self, invoice_df):
        extracted_data = {
            'contractor': {'value': "", 'cell': ""},
            'number': {'value': "", 'cell': ""},
            'date': {'value': None, 'cell': ""},
            'items': []
        }

        # Extract contractor name
        contractor_cell_location = self.config.get('contractor_cell', DEFAULT_CONFIG['contractor_cell'])
        contractor_name_cell = self.get_cell_value(invoice_df, contractor_cell_location)
        if contractor_name_cell:
            # Extract text between quotes if present
            quotes_match = re.search(r'"([^"]+)"', str(contractor_name_cell))
            if quotes_match:
                extracted_data['contractor']['value'] = quotes_match.group(1).strip()
            else:
                extracted_data['contractor']['value'] = str(contractor_name_cell).strip()
            extracted_data['contractor']['cell'] = contractor_cell_location
        self.log_message(f"Извлечен контрагент: {extracted_data['contractor']['value']} (ячейка: {contractor_cell_location})")

        # Extract invoice number
        invoice_number_cell_location = self.config.get('number_cell', DEFAULT_CONFIG['number_cell'])
        invoice_number_cell = self.get_cell_value(invoice_df, invoice_number_cell_location)
        if invoice_number_cell:
            extracted_data['number']['value'] = str(invoice_number_cell).strip()
            extracted_data['number']['cell'] = invoice_number_cell_location
        self.log_message(f"Извлечен номер счета-фактуры: {extracted_data['number']['value']} (ячейка: {invoice_number_cell_location})")

        # Extract date
        invoice_date_cell_location = self.config.get('date_cell', DEFAULT_CONFIG['date_cell'])
        invoice_date_str_cell = self.get_cell_value(invoice_df, invoice_date_cell_location)
        if invoice_date_str_cell:
            date_str = str(invoice_date_str_cell).strip()
            try:
                extracted_data['date']['value'] = datetime.strptime(date_str, '%d.%m.%Y').date()
            except ValueError:
                try:
                    extracted_data['date']['value'] = datetime.strptime(date_str, '%d %B %Y').date()
                except ValueError:
                    try:
                        extracted_data['date']['value'] = datetime.strptime(date_str, '%d %B %Y г.').date()
                    except ValueError:
                        extracted_data['date']['value'] = date_str
                        self.log_message(f"Предупреждение: Формат даты не распознан в ячейке {invoice_date_cell_location}. Данные сохранены как текст.")
        extracted_data['date']['cell'] = invoice_date_cell_location
        self.log_message(f"Извлечена дата счета-фактуры: {extracted_data['date']['value']} (ячейка: {invoice_date_cell_location})")

        # Extract items
        items_data = []
        start_row_index = None
        items_start_cell_location = self.config.get('items_start_cell', DEFAULT_CONFIG['items_start_cell'])
        start_row_index, start_col_index = self.excel_cell_to_index(items_start_cell_location)

        if start_row_index is not None:
            item_text_part_col_letter = self.config.get('items_cell_text_part', DEFAULT_CONFIG['items_cell_text_part'])
            item_numeric_part_col_letter = self.config.get('items_cell_numeric_part', DEFAULT_CONFIG['items_cell_numeric_part'])
            item_weight_col_letter = self.config.get('items_cell_weight', DEFAULT_CONFIG['items_cell_weight'])
            item_price_col_letter = self.config.get('items_cell_price', DEFAULT_CONFIG['items_cell_price'])

            row_offset = start_row_index
            item_index = 0
            while True:
                current_row_index = row_offset + item_index

                item_name_full_cell_location = f"{item_text_part_col_letter}{current_row_index + 1}"
                item_name_full = self.get_cell_value(invoice_df, item_name_full_cell_location)

                if not item_name_full:
                    break

                self.log_message(f"Чтение позиции из ячейки {item_name_full_cell_location}: Значение = '{item_name_full}'")

                item_weight_cell_location = f"{item_weight_col_letter}{current_row_index + 1}"
                item_weight = self.get_cell_value(invoice_df, item_weight_cell_location)
                item_price_cell_location = f"{item_price_col_letter}{current_row_index + 1}"
                item_price = self.get_cell_value(invoice_df, item_price_cell_location)

                item_text_part = ""
                item_numeric_part = ""
                parts = re.split(r'(\d+)', str(item_name_full), 1)
                item_text_part = parts[0].strip()
                if len(parts) > 1:
                    item_numeric_part = parts[1].strip() + "".join(parts[2:]).strip()

                items_data.append({
                    'text_part': {'value': item_text_part, 'cell': item_name_full_cell_location},
                    'numeric_part': {'value': item_numeric_part, 'cell': item_name_full_cell_location},
                    'weight': {'value': item_weight, 'cell': item_weight_cell_location},
                    'price': {'value': item_price, 'cell': item_price_cell_location},
                })
                item_index += 1

        self.log_message(f"Извлечено позиций: {len(items_data)}")
        extracted_data['items'] = items_data
        return extracted_data

    def save_with_formatting(self, new_df):
        try:
            if not Path(self.output_file).exists():
                with pd.ExcelWriter(self.output_file, engine='openpyxl') as writer:
                    new_df.to_excel(writer, sheet_name='Sheet1', index=False, startrow=4)
                return

            # Load existing workbook
            book = openpyxl.load_workbook(self.output_file)
            sheet = book.active

            # Read existing data for combining
            existing_df = pd.read_excel(self.output_file, skiprows=4, header=None)
            existing_df = existing_df.dropna(how='all')

            # Combine data
            combined_df = pd.concat([existing_df, new_df], ignore_index=True)
            combined_df = combined_df.dropna(how='all')

            # Store original formatting
            cell_formats = {}
            for row in range(5, sheet.max_row + 1):
                for col in range(1, sheet.max_column + 1):
                    cell = sheet.cell(row=row, column=col)
                    cell_formats[(row, col)] = {
                        'number_format': cell.number_format,
                        'font': cell.font.copy(),
                        'alignment': cell.alignment.copy(),
                        'border': cell.border.copy(),
                        'fill': cell.fill.copy()
                    }

            # Clear existing content while preserving header rows
            for row in range(5, sheet.max_row + 1):
                for col in range(1, sheet.max_column + 1):
                    sheet.cell(row=row, column=col).value = None

            # Write new data
            for index, row in combined_df.iterrows():
                for col_index, value in enumerate(row):
                    if pd.notna(value):
                        cell = sheet.cell(row=index + 5, column=col_index + 1)
                        cell.value = value
                        
                        # Restore formatting if available
                        if (index + 5, col_index + 1) in cell_formats:
                            fmt = cell_formats[(index + 5, col_index + 1)]
                            cell.number_format = fmt['number_format']
                            cell.font = fmt['font']
                            cell.alignment = fmt['alignment']
                            cell.border = fmt['border']
                            cell.fill = fmt['fill']

            book.save(self.output_file)
            self.log_message("Данные успешно сохранены с сохранением форматирования")

        except Exception as e:
            self.log_message(f"Ошибка при сохранении файла: {str(e)}")
            raise

    def process_invoice(self):
        if not self.output_file:
            messagebox.showerror("Ошибка", "Сначала выберите файл отчета")
            return

        invoice_file = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if not invoice_file:
            return

        try:
            invoice_df = pd.read_excel(invoice_file, header=None)
            extracted_data_with_cells = self.extract_invoice_data(invoice_df)
            dialog = DataReviewDialog(self.root, extracted_data_with_cells, self.config)
            result_data, updated_config = dialog.result

            if result_data:
                self.config = updated_config
                self.save_config()

                # Create DataFrame for new invoice data only
                new_rows = []
                for i, item in enumerate(result_data['items']):
                    row = {
                        0: result_data['number']['value'] if i == 0 else '',  # Changed from 'C1' to 0
                        1: result_data['contractor']['value'] if i == 0 else '',  # Changed from 'C2' to 1
                        2: result_data['date']['value'].strftime('%d.%m.%Y') if isinstance(result_data['date']['value'], datetime) and i == 0 else (str(result_data['date']['value']) if i == 0 else ''),
                        3: 'Э' if i == 0 else '',
                        4: item['text_part']['value'],
                        5: '',
                        6: item['numeric_part']['value'],
                        7: item['weight']['value'],
                        8: item['price']['value'],
                        9: '',
                        10: '',
                        11: f"{result_data['number']['value']} от {result_data['date']['value'].strftime('%d.%m.%Y')}" if isinstance(result_data['date']['value'], datetime) and i == 0 else (f"{result_data['number']['value']} от {str(result_data['date']['value'])}" if i == 0 else '')
                    }
                    new_rows.append(row)

                new_df = pd.DataFrame(new_rows)
                self.save_with_formatting(new_df)
                self.log_message(f"Invoice processed successfully: {invoice_file}")
            else:
                self.log_message("Invoice processing cancelled by user.")

        except Exception as e:
            messagebox.showerror("Ошибка", f"Ошибка обработки файла: {str(e)}")
            self.log_message(f"Error processing file: {str(e)}")

    def get_cell_value(self, df, cell_location):
        """Get value from cell in DataFrame"""
        try:
            row_index, col_index = self.excel_cell_to_index(cell_location)
            if row_index >= 0 and col_index >= 0:
                cell_value = df.iloc[row_index, col_index]
                return str(cell_value) if not pd.isna(cell_value) else ""
            return ""
        except (IndexError, ValueError, TypeError) as e:
            self.log_message(f"Ошибка при чтении ячейки {cell_location}: {e}")
            return ""

    def get_last_number(self, output_df):
        """
        Получает последний номер строки для вставки новых данных.
        Ищет строку для добавления новой счет-фактуры в столбце 4 (колонка E, ранее C5).
        Если файл пустой или колонка 4 не найдена, возвращает 0.
        """
        self.log_message("Вызвана функция get_last_number")
        self.log_message(f"Колонны в output_df: {list(output_df.columns)}") # List для читабельности

        if output_df.empty or 4 not in output_df.columns:
            self.log_message("Файл отчета пуст или нет колонки 4, возвращаем 0")
            return 0

        c4_column = output_df[4].dropna()
        self.log_message(f"Колонка 4 (dropna): {c4_column}")

        if c4_column.empty:
            self.log_message("Колонка 4 пуста после dropna, возвращаем 0")
            return 0

        last_index = c4_column.last_valid_index()
        self.log_message(f"last_valid_index: {last_index}")

        if last_index is None:
            self.log_message("last_index is None, возвращаем 0")
            return 0

        next_row_index = last_index # Возвращаем индекс последней строки, а не следующую за ней, т.к. save_with_formatting добавляет +5 и +1
        self.log_message(f"Последний индекс найден: {last_index}, строка для начала записи: {next_row_index + 1} (Excel нумерация)") # +1 для Excel нумерации
        return next_row_index # Возвращаем индекс последней строки

    def load_config(self):
        try:
            with open(CONFIG_FILE, 'r') as f:
                config = json.load(f)
                self.log_message(f"Конфигурация загружена из файла: {CONFIG_FILE}")
                return config
        except FileNotFoundError:
            self.log_message(f"Файл конфигурации не найден: {CONFIG_FILE}. Используется конфигурация по умолчанию. Файл будет создан при сохранении.")
            return DEFAULT_CONFIG.copy()
        except json.JSONDecodeError:
            self.log_message(f"Файл конфигурации поврежден: {CONFIG_FILE}. Конфигурация сброшена к значениям по умолчанию.")
            return DEFAULT_CONFIG.copy()
        except Exception as e:
            self.log_message(f"Непредвиденная ошибка при загрузке конфигурации: {str(e)}. Используется конфигурация по умолчанию.")
            return DEFAULT_CONFIG.copy()

    def save_config(self):
        try:
            with open(CONFIG_FILE, 'w') as f:
                json.dump(self.config, f, indent=4)
                self.log_message(f"Конфигурация успешно сохранена в файл: {CONFIG_FILE}")
        except Exception as e:
            self.log_message(f"Ошибка сохранения конфигурации в файл: {CONFIG_FILE}. Ошибка: {e}")

    def save_config(self):
        try:
            config_dir = os.path.dirname(os.path.abspath(__file__))
            config_path = os.path.join(config_dir, CONFIG_FILE)
            with open(config_path, 'w') as f:
                json.dump(self.config, f, indent=4)
                self.log_message(f"Конфигурация успешно сохранена в файл: {config_path}")
        except Exception as e:
            self.log_message(f"Ошибка сохранения конфигурации в файл: {CONFIG_FILE}. Ошибка: {e}")

    def excel_cell_to_index(self, cell_location):
        """Convert Excel cell reference to DataFrame indices"""
        try:
            if 'R' in cell_location and 'C' in cell_location:
                parts = cell_location.upper().split('C')
                row_index = int(parts[0].replace('R', '')) - 1
                col_index = int(parts[1]) - 1
            else:
                column_letter = ''.join(filter(str.isalpha, cell_location.upper()))
                row_number = int(''.join(filter(str.isdigit, cell_location)))
                col_index = 0
                for letter in column_letter:
                    col_index = col_index * 26 + (ord(letter) - ord('A') + 1)
                row_index = row_number - 1
                col_index -= 1
            return row_index, col_index
        except Exception as e:
            self.log_message(f"Ошибка при преобразовании адреса ячейки {cell_location}: {e}")
            return None, None

    def index_to_excel_cell(self, row_index, col_index):
        col_letter = ''
        while col_index >= 0:
            col_index, remainder = divmod(col_index, 26)
            col_letter = chr(ord('A') + remainder) + col_letter
            if col_index == 0:
                break
            col_index -= 1
        return f"{col_letter}{row_index + 1}"


class DataReviewDialog(simpledialog.Dialog):
    def __init__(self, parent, data, config):
        self.extracted_data = data
        self.config = config
        self.result = None
        super().__init__(parent, "Проверка и редактирование данных")

    def body(self, frame):
        tk.Label(frame, text="Проверьте и отредактируйте данные, при необходимости укажите ячейки вручную:").pack(anchor=tk.W)
        self.entries = {}

        self.create_data_row(frame, 'contractor', "Контрагент:")
        self.create_data_row(frame, 'number', "Номер счета-фактуры:")
        self.create_data_row(frame, 'date', "Дата счета-фактуры:")

        tk.Label(frame, text="Позиции:").pack(anchor=tk.W)
        items_frame = tk.Frame(frame, bd=1, relief=tk.SOLID, padx=5, pady=5)
        items_frame.pack(fill=tk.X, expand=True, anchor=tk.W, padx=10, pady=5)

        for i, item in enumerate(self.extracted_data['items']):
            item_label = tk.Label(items_frame, text=f"Позиция {i+1}:", anchor=tk.W)
            item_label.grid(row=i, column=0, sticky=tk.W)

            self.create_item_row(items_frame, item, 'text_part', i, 1, "Текст:")
            self.create_item_row(items_frame, item, 'numeric_part', i, 2, "Цифра:")
            self.create_item_row(items_frame, item, 'weight', i, 3, "Вес:")
            self.create_item_row(items_frame, item, 'price', i, 4, "Цена:")

        return frame

    def create_data_row(self, frame, data_key, label_text):
        data_frame = tk.Frame(frame)
        data_frame.pack(fill=tk.X, expand=True, anchor=tk.W, padx=20)

        tk.Label(data_frame, text=label_text).pack(side=tk.LEFT, anchor=tk.W)
        value_label = tk.Label(data_frame, text=f"Значение: {self.extracted_data[data_key]['value']}", anchor=tk.W)
        value_label.pack(side=tk.LEFT)
        cell_label = tk.Label(data_frame, text=f"Ячейка: {self.extracted_data[data_key]['cell']}", anchor=tk.W, padx=10)
        cell_label.pack(side=tk.LEFT)

        entry = tk.Entry(data_frame, width=10)
        entry.pack(side=tk.LEFT, padx=10)
        entry.insert(0, self.config.get(f"{data_key}_cell", ""))
        self.entries[data_key] = entry

    def create_item_row(self, frame, item, item_key, row_index, col_index, label_text):
        item_frame = tk.Frame(frame)
        item_frame.grid(row=row_index, column=col_index, sticky=tk.W + tk.E + tk.N + tk.S, padx=5)

        tk.Label(item_frame, text=label_text, anchor=tk.W).pack(anchor=tk.W)
        value_label = tk.Label(item_frame, text=f"Значение: {item[item_key]['value']}", anchor=tk.W, wraplength=150, justify=tk.LEFT)
        value_label.pack(anchor=tk.W)
        cell_label = tk.Label(item_frame, text=f"Ячейка: {item[item_key]['cell']}", anchor=tk.W)
        cell_label.pack(anchor=tk.W)
        entry = tk.Entry(item_frame, width=8)
        entry.pack(anchor=tk.W)
        config_key = f"items_cell_{item_key}"
        entry.insert(0, self.config.get(config_key, ""))
        self.entries[f"items_{row_index}_{item_key}"] = entry

    def apply(self):
        updated_config = self.config.copy()
        updated_data = self.extracted_data.copy()

        for key in ['contractor', 'number', 'date']:
            cell_entry = self.entries[key].get().strip().upper()
            if cell_entry:
                updated_config[f"{key}_cell"] = cell_entry
                updated_data[key]['cell'] = cell_entry

        for i, item in enumerate(updated_data['items']):
            for item_key in ['text_part', 'numeric_part', 'weight', 'price']:
                cell_entry = self.entries[f"items_{i}_{item_key}"].get().strip().upper()
                if cell_entry:
                    config_key = f"items_cell_{item_key}"
                    updated_config[config_key] = cell_entry
                    item[item_key]['cell'] = cell_entry

        self.result = updated_data, updated_config
        self.config = updated_config
        self.destroy()

    def cancel_command(self):
        self.result = None
        super().cancel_command()


if __name__ == "__main__":
    root = tk.Tk()
    app = InvoiceProcessor(root)
    root.mainloop()